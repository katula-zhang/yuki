#!/usr/bin/python3
from openpyxl import load_workbook
from openpyxl import workbook
import os,sys,struct,string
DIR = os.path.abspath('.')
sourceDir = DIR + '\\sourceFile\\'
objectDir = DIR + '\\objectFile\\'
tmpFile1 = '部门考核模板'
tmpFile2 = '个人考核模板'

def copyResult1(str):#个人
        global tCount
        global result_Wb
        global result_Ws
        global tName
        wb = load_workbook(str)
        ws = wb.get_sheet_by_name('Sheet1')
        rst = []
        for i in range(1, 13):
                #print(ws["F%d" % i].value)
                rst.append(ws["F%d" % i].value)
        rst.append(ws['c14'].value)
        print(rst)
        rst[2] = rName
        print(rst)
        for i in range(1, 12):
                result_Ws.cell(row = i + 1, column = tCount, value = rst[i])
        result_Ws.cell(row = 14, column = tCount, value = rst[12])
        tCount += 1
        wb.close()
def copyResult2(str):#部门
        global tCount
        global result_Wb
        global result_Ws
        global tName
        wb = load_workbook(str)
        ws = wb.get_sheet_by_name('Sheet1')
        rst = []
        for i in range(1, 11):
                #print(ws["F%d" % i].value)
                rst.append(ws["E%d" % i].value)
        rst.append(ws['c12'].value)
        print(rst)
        rst[2] = rName
        print(rst)
        for i in range(1, 10):
                result_Ws.cell(row = i + 1, column = tCount, value = rst[i])
        result_Ws.cell(row = 12, column = tCount, value = rst[10])
        tCount += 1
        wb.close()

files = os.listdir(sourceDir)
lastName = "初值"
lastPartName = ""
tName = " " #被考评人
rName = " " #打分人
tCount = 7
#result_Wb
#result_Ws
for file in files :
        sub = file.split('-')
        print(sub)
        if(len(sub) > 2):
                tName = sub[1]
                rName = sub[2]
                if(lastName == '初值'):
                        result_Wb = load_workbook(tmpFile2 + '.xlsx')
                        result_Ws = result_Wb.get_sheet_by_name('Sheet1')
                elif(lastName != tName):
                        result_Wb.save(objectDir + lastPartName + "_"+ lastName + ".xlsx")
                        result_Wb.close()
                        tCount = 7
                        result_Wb = load_workbook(tmpFile2 + '.xlsx')
                        result_Ws = result_Wb.get_sheet_by_name('Sheet1')
                lastPartName = sub[0]
                copyResult1(sourceDir + "\\" + file)
                lastName = tName
        else:
                tName = sub[0]
                rName = sub[1]
                if(lastName == "初值"):
                        result_Wb = load_workbook(tmpFile1 + '.xlsx')
                        result_Ws = result_Wb.get_sheet_by_name('Sheet1')
                elif(lastName != tName):
                        result_Wb.save(objectDir + lastPartName + "_"+ lastName + ".xlsx")
                        result_Wb.close()
                        tCount = 7
                        result_Wb = load_workbook(tmpFile1 + '.xlsx')
                        result_Ws = result_Wb.get_sheet_by_name('Sheet1')
                lastPartName = sub[0]
                copyResult2(sourceDir + "\\" + file)
                lastName = tName
result_Wb.save(objectDir + lastPartName + "_"+ lastName + ".xlsx")
result_Wb.close()        
