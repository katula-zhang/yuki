#!/usr/bin/python3
import os,sys,struct,string
from openpyxl import load_workbook
from openpyxl import workbook
DIR = os.path.abspath('.')
binDIR = DIR + '\\' + "ui_config.xlsx"
wb = load_workbook(binDIR)
sheet1 = wb.get_sheet_by_name('config')
print(sheet1.title)
configdata = []
configdata.append(int(sheet1['C2'].value))
configdata.append(int(sheet1['C3'].value))
configdata.append(int(sheet1['C4'].value))
start_row = 6
ws_rows_len = sheet1.max_row
print("max_column:%d" % ws_rows_len)
start_col = 8
for row in range(start_row, ws_rows_len + 1):
        for column in range(3,9):
                t = sheet1.cell(row=row,column=column).value
                #print(t)
                configdata.append(t)
l = len(configdata)
if(l < 200):
        for i in range(200 - l):
                configdata.append(0)
outAddr = DIR + '\\' + "bmpaddr.bin"
addrFile = open(outAddr, "ab")
for ad in configdata:
        ta = struct.pack('B', ad)
        addrFile.write(ta)




